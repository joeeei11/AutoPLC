"""Tests for Excel I/O table exporter."""

from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from plc_logic_gen.exporter_excel import _HEADERS, export_io_excel
from plc_logic_gen.models.ld import IOSignal, SignalType


def _di(tag: str, name: str = "") -> IOSignal:
    return IOSignal(tag=tag, name=name, signal_type=SignalType.DI)


def _ai(tag: str, name: str = "", low: float = 0.0, high: float = 100.0, unit: str = "") -> IOSignal:
    return IOSignal(tag=tag, name=name, signal_type=SignalType.AI, range_low=low, range_high=high, engineering_unit=unit)


class TestExportIOExcel:
    def test_returns_nonempty_bytes(self):
        result = export_io_excel([_di("DI001")])
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_parseable_by_openpyxl(self):
        result = export_io_excel([_di("DI001")])
        wb = load_workbook(io.BytesIO(result))
        assert wb is not None

    def test_correct_column_count(self):
        result = export_io_excel([_di("DI001")])
        ws = load_workbook(io.BytesIO(result)).active
        assert ws.max_column == 10

    def test_header_row_values(self):
        result = export_io_excel([_di("DI001")])
        ws = load_workbook(io.BytesIO(result)).active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 11)]
        assert headers == _HEADERS

    def test_row_count_matches_signals(self):
        signals = [_di("DI001"), _di("DI002"), _ai("AI_TT101")]
        result = export_io_excel(signals)
        ws = load_workbook(io.BytesIO(result)).active
        # header row + data rows
        assert ws.max_row == 1 + len(signals)

    def test_tag_value_in_first_data_row(self):
        result = export_io_excel([_di("DI001", "启动按钮")])
        ws = load_workbook(io.BytesIO(result)).active
        assert ws.cell(row=2, column=1).value == "DI001"
        assert ws.cell(row=2, column=2).value == "启动按钮"

    def test_di_signal_range_fields_are_none(self):
        result = export_io_excel([_di("DI001")])
        ws = load_workbook(io.BytesIO(result)).active
        assert ws.cell(row=2, column=7).value is None  # range_low
        assert ws.cell(row=2, column=8).value is None  # range_high
        assert ws.cell(row=2, column=9).value == ""    # engineering_unit

    def test_ai_signal_range_fields_present(self):
        result = export_io_excel([_ai("AI_TT101", low=0.0, high=200.0, unit="℃")])
        ws = load_workbook(io.BytesIO(result)).active
        assert ws.cell(row=2, column=7).value == 0.0
        assert ws.cell(row=2, column=8).value == 200.0
        assert ws.cell(row=2, column=9).value == "℃"

    def test_empty_signals_produces_header_only(self):
        result = export_io_excel([])
        ws = load_workbook(io.BytesIO(result)).active
        assert ws.max_row == 1

    def test_header_cell_fill_is_blue(self):
        result = export_io_excel([_di("DI001")])
        ws = load_workbook(io.BytesIO(result), data_only=True).active
        fill = ws.cell(row=1, column=1).fill
        assert fill.fgColor.rgb == "FF4472C4"

    def test_header_font_is_bold_white(self):
        result = export_io_excel([_di("DI001")])
        ws = load_workbook(io.BytesIO(result)).active
        font = ws.cell(row=1, column=1).font
        assert font.bold is True
        assert font.color.rgb == "FFFFFFFF"
