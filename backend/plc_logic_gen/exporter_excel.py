"""Excel exporter for IOSignal lists using openpyxl."""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from plc_logic_gen.models.ld import IOSignal, SignalType

_HEADERS = [
    "Tag", "信号名称", "信号类型", "PLC地址",
    "模块号", "通道号", "量程低", "量程高", "工程单位", "注释",
]

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ALT_FILL = PatternFill("solid", fgColor="DEEAF1")


def export_io_excel(signals: list[IOSignal]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "I/O分配表"

    ws.append(_HEADERS)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for i, sig in enumerate(signals, start=2):
        is_analog = sig.signal_type in (SignalType.AI, SignalType.AO)
        row = [
            sig.tag,
            sig.name,
            sig.signal_type.value,
            sig.plc_address,
            sig.module_no,
            sig.channel_no,
            sig.range_low if is_analog else None,
            sig.range_high if is_analog else None,
            sig.engineering_unit if is_analog else "",
            sig.comment,
        ]
        ws.append(row)
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = _ALT_FILL

    # Auto-fit column widths (12–40)
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
            max(12, min(40, max_len + 2))
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
